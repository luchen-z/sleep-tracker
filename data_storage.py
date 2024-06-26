import json
import os
import pprint
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

from utils import sort_nested_dict

SLEEP_DATA_DIRECTORY = 'data_store/sleep_data'
GENERATED_DATA_DIRECTORY = 'data_store/generated'
CONFIG_FILE = 'data_store/user_config.json'


# Function to read all user JSON files and combine them into a single dictionary
def read_json_file():
    directory = SLEEP_DATA_DIRECTORY
    if not os.path.exists(directory):
        os.makedirs(directory)

    all_data = {}

    for filename in os.listdir(directory):
        if filename.endswith(".json"):
            user = filename[:-5]  # Remove the .json extension to get the user name
            filepath = os.path.join(directory, filename)
            with open(filepath) as fp:
                data_loaded = json.load(fp)
                all_data[user] = sort_nested_dict(data_loaded)

    pprint.pp(all_data)
    return all_data


def save_to_files(users_records):
    save_to_json_file(users_records)
    save_to_excel_file(users_records)
    return users_records


# Function to save user records to separate JSON files
def save_to_json_file(users_records):
    directory = SLEEP_DATA_DIRECTORY
    if not os.path.exists(directory):
        os.makedirs(directory)

    users_records = sort_nested_dict(users_records)

    for user, records in users_records.items():
        filename = os.path.join(directory, f"{user}.json")
        with open(filename, 'w') as f:
            json.dump(records, f, indent=4)
    return users_records


# Function to save DataFrame to Excel with formatting
def save_to_excel_file(users_records):
    output_filename = os.path.join(GENERATED_DATA_DIRECTORY, 'sleep_records.xlsx')

    # Create a Pandas Excel writer using openpyxl as the engine
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        for user, all_records in users_records.items():
            # Group the records by month and year
            records_by_month = {}
            for date, entries in all_records.items():
                month_year = pd.to_datetime(date).strftime('%Y-%m')
                if month_year not in records_by_month:
                    records_by_month[month_year] = {}
                records_by_month[month_year][date] = entries

            for month_year, records in records_by_month.items():
                # Get the first and last day of the month
                first_day = pd.to_datetime(f'{month_year}-01')
                last_day = first_day + pd.DateOffset(months=1) - pd.DateOffset(days=1)
                days = pd.date_range(start=first_day, end=last_day).strftime('%m/%d')

                # Create an empty DataFrame to store the sleep data with float data type
                df = pd.DataFrame(0.0, index=range(24), columns=days)  # Hours from 0 to 23

                # Populate the DataFrame with sleep data
                for date, entries in records.items():
                    formatted_date = pd.to_datetime(date).strftime('%m/%d')
                    for entry in entries:
                        start_time = pd.to_datetime(entry[0], format='%H:%M')
                        end_time = pd.to_datetime(entry[1], format='%H:%M')

                        current_time = start_time
                        while current_time < end_time:
                            hour = current_time.hour
                            next_hour = (current_time + pd.Timedelta(hours=1)).replace(minute=0, second=0,
                                                                                       microsecond=0)

                            if next_hour <= end_time:
                                remaining_time = (next_hour - current_time).total_seconds() / 3600
                                df.at[hour, formatted_date] += round(remaining_time, 2)
                                current_time = next_hour
                            else:
                                remaining_time = (end_time - current_time).total_seconds() / 3600
                                df.at[hour, formatted_date] += round(remaining_time, 2)
                                break

                # Clear cells with a value of 0
                df.replace(0, None, inplace=True)

                # Write DataFrame to Excel
                df.to_excel(writer, sheet_name=f'{user}_{month_year}')

                # Get the workbook and the worksheet
                worksheet = writer.sheets[f'{user}_{month_year}']

                # Apply formatting
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2,
                                               max_col=worksheet.max_column):
                    for cell in row:
                        if cell.value:
                            cell.fill = PatternFill(start_color='FFCC00', end_color='FFCC00', fill_type='solid')
                            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                 top=Side(style='thin'), bottom=Side(style='thin'))
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.font = Font(bold=True)
                        else:
                            cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

                # Adjust column widths
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter  # Get the column name
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column].width = adjusted_width


def read_user_name():
    if not os.path.exists(CONFIG_FILE):
        return None
    with open(CONFIG_FILE, 'r') as file:
        config = json.load(file)
    return config.get('user_name', None)


def save_user_name(user_name):
    config = {"user_name": user_name}
    with open(CONFIG_FILE, 'w') as file:
        json.dump(config, file, indent=4)
